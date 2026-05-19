#!/usr/bin/env python3
"""Enter January 2026 match scores from Excel into event_id=23 via API."""

import csv
import json
import subprocess
import openpyxl

EXCEL = '/Users/Kalibek_Turgumbayev/projects/test/league/docs/Лиги Астана 2026_01_11.xlsx'
USERS_CSV = '/tmp/users.csv'
COOKIES = '/tmp/admin_cookies.txt'
BASE_URL = 'http://localhost:8080/api/v1'

GROUP_IDS = {
    ('S', 1): 45,
    ('A', 1): 46, ('A', 2): 47, ('A', 3): 48, ('A', 4): 49,
    ('A', 5): 50, ('A', 6): 51, ('A', 7): 52, ('A', 8): 53,
    ('B', 1): 56, ('B', 2): 57, ('B', 3): 58, ('B', 4): 59,
    ('B', 5): 60, ('B', 6): 61, ('B', 7): 62, ('B', 8): 63,
    ('C', 1): 64, ('C', 2): 65, ('C', 3): 66, ('C', 4): 67,
    ('C', 5): 68, ('C', 6): 69, ('C', 7): 70, ('C', 8): 71,
    ('D', 1): 72, ('D', 2): 73, ('D', 3): 74, ('D', 4): 75,
    ('D', 5): 76, ('D', 6): 77, ('D', 7): 78, ('D', 8): 79,
    ('E', 1): 80, ('E', 2): 81, ('E', 3): 82, ('E', 4): 83,
    ('E', 5): 84, ('E', 6): 85, ('E', 7): 86, ('E', 8): 87,
}

CYRILLIC_MAP = {'А': 'A', 'В': 'B', 'С': 'C', 'Д': 'D', 'Е': 'E'}

def normalize_division(ch):
    return CYRILLIC_MAP.get(ch, ch).upper()

def parse_sheet_name(name):
    if name in ('Суперлига', 'Superleague'):
        return ('S', 1)
    base = name.replace('_2', '') if name.endswith('_2') else name
    div = normalize_division(base[0])
    try:
        no = int(base[1:])
    except ValueError:
        return None
    return (div, no)

def load_users():
    users = {}
    with open(USERS_CSV, newline='', encoding='utf-8') as f:
        for row in csv.reader(f):
            if len(row) < 3:
                continue
            try:
                uid = int(row[0].strip())
            except ValueError:
                continue
            first, last = row[1].strip(), row[2].strip()
            users[(last.upper(), first.upper())] = uid
    # Spelling-variant aliases for name mismatches between player list and match rows:
    users[('ШИПИЦЫНА', 'ДАРЬЯ')] = 389   # Шипицина/Шипицына same person; group has 389
    users[('СЕМЕНОВА', 'ЯНА')] = 412      # Сёменова/Семенова same person; group has 412
    users[('ДУСЕНБАЕВ', 'МУРАТ')] = 183   # Дюсенбаев/Дусенбаев same person; group has 183
    users[('ЕРЕМИН', 'ИВАН')] = 197       # Ерёмин/Еремин (ё/е) same person; group has 197
    return users

def find_uid(full_name, users):
    parts = full_name.strip().split(None, 1)
    if len(parts) == 2:
        last, first = parts[0].upper(), parts[1].upper()
        if (last, first) in users:
            return users[(last, first)]
        if (first, last) in users:
            return users[(first, last)]
    return None

def get_group_data(group_id):
    r = subprocess.run(
        ['curl', '-s', f'{BASE_URL}/public/events/23/groups/{group_id}'],
        capture_output=True, text=True
    )
    return json.loads(r.stdout)

def put_score(group_id, match_id, score1, score2):
    r = subprocess.run(
        ['curl', '-s', '-b', COOKIES, '-X', 'PUT',
         f'{BASE_URL}/secured/groups/{group_id}/matches/{match_id}',
         '-H', 'Content-Type: application/json',
         '-d', json.dumps({'score1': score1, 'score2': score2})],
        capture_output=True, text=True
    )
    return r.stdout.strip()

def parse_score(score_str):
    """Parse '3:1' → (3, 1). Returns None if unparseable."""
    if not score_str or not isinstance(score_str, str):
        return None
    # Handle W:L / L:W forfeit scores
    s = score_str.strip()
    if ':' not in s:
        return None
    parts = s.split(':')
    if len(parts) != 2:
        return None
    # Skip W/L forfeit entries — these need withdraw flags, handle separately
    if parts[0].isalpha() or parts[1].isalpha():
        return None
    try:
        return int(parts[0]), int(parts[1])
    except ValueError:
        return None

def extract_matches(ws):
    """Extract (player1_name, player2_name, score1, score2) from match rows.
    Returns list of tuples; score1/score2 are None if match not played."""
    matches = []
    for row in ws.iter_rows(min_row=30, max_row=80, values_only=True):
        p1 = row[9]   # col J
        sep = row[10] # col K
        p2 = row[11]  # col L
        score_raw = row[15]  # col P
        if not (isinstance(p1, str) and sep == '-' and isinstance(p2, str)):
            continue
        p1 = p1.strip().rstrip('+')
        p2 = p2.strip().rstrip('+')
        score = parse_score(score_raw)
        s1, s2 = (score[0], score[1]) if score else (None, None)
        matches.append((p1, p2, s1, s2))
    return matches

def process_group(sheet_name, group_key, users):
    group_id = GROUP_IDS.get(group_key)
    if group_id is None:
        return

    # Get API data
    data = get_group_data(group_id)
    players = data.get('players', [])
    api_matches = data.get('matches', [])

    # Build userId → groupPlayerId
    uid_to_gpid = {p['userId']: p['groupPlayerId'] for p in players}

    # Build {frozenset(gp1,gp2) → (matchId, gp1, gp2)}
    match_lookup = {}
    for m in api_matches:
        gp1, gp2 = m['groupPlayer1Id'], m['groupPlayer2Id']
        key = frozenset([gp1, gp2])
        match_lookup[key] = (m['matchId'], gp1, gp2)

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb[sheet_name]
    excel_matches = extract_matches(ws)

    div, no = group_key
    print(f"\n--- {div}{no} (groupId={group_id}, sheet={sheet_name}) ---")

    ok_count = no_score = not_found = error_count = 0

    for p1_name, p2_name, s1, s2 in excel_matches:
        uid1 = find_uid(p1_name, users)
        uid2 = find_uid(p2_name, users)

        if uid1 is None or uid2 is None:
            missing = p1_name if uid1 is None else p2_name
            print(f"  USER NOT FOUND: {missing!r}")
            not_found += 1
            continue

        gp1 = uid_to_gpid.get(uid1)
        gp2 = uid_to_gpid.get(uid2)

        if gp1 is None or gp2 is None:
            missing = p1_name if gp1 is None else p2_name
            print(f"  NOT IN GROUP: {missing!r} (uid={uid1 if gp1 is None else uid2})")
            not_found += 1
            continue

        if s1 is None:
            # No score — skip (DNS or unplayed)
            no_score += 1
            continue

        key = frozenset([gp1, gp2])
        if key not in match_lookup:
            print(f"  MATCH NOT FOUND: {p1_name} vs {p2_name}")
            not_found += 1
            continue

        match_id, api_gp1, api_gp2 = match_lookup[key]
        # Align score to API player order
        if api_gp1 == gp1:
            final_s1, final_s2 = s1, s2
        else:
            final_s1, final_s2 = s2, s1

        resp = put_score(group_id, match_id, final_s1, final_s2)
        try:
            data_r = json.loads(resp)
            if data_r.get('ok'):
                ok_count += 1
            else:
                print(f"  ERROR: {p1_name} vs {p2_name} {s1}:{s2} → {resp}")
                error_count += 1
        except json.JSONDecodeError:
            print(f"  RAW: {p1_name} vs {p2_name}: {resp}")
            error_count += 1

    print(f"  Scored: {ok_count} | Skipped(no score): {no_score} | Errors: {error_count} | Not found: {not_found}")
    return ok_count, no_score, error_count, not_found

def main():
    users = load_users()
    print(f"Loaded {len(users)} users")

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    sheets_to_process = []
    for sname in wb.sheetnames:
        key = parse_sheet_name(sname)
        if key and GROUP_IDS.get(key):
            # Check if sheet has any match data
            ws = wb[sname]
            has_matches = any(
                isinstance(row[9], str) and row[10] == '-' and isinstance(row[11], str)
                for row in ws.iter_rows(min_row=30, max_row=80, values_only=True)
            )
            if has_matches:
                sheets_to_process.append((sname, key))

    # Deduplicate: if both primary and _2 exist, use whichever has match data
    seen_keys = {}
    for sname, key in sheets_to_process:
        if key not in seen_keys:
            seen_keys[key] = sname
        else:
            # Prefer the one with more matches (non-_2 vs _2)
            # Keep both to check — but since we validated has_matches, just keep first
            pass

    total_ok = total_skip = total_err = total_nf = 0
    for key, sname in seen_keys.items():
        result = process_group(sname, key, users)
        if result:
            ok, skip, err, nf = result
            total_ok += ok
            total_skip += skip
            total_err += err
            total_nf += nf

    print(f"\n=== TOTAL SUMMARY ===")
    print(f"Scored: {total_ok} | Skipped(no score/DNS): {total_skip} | Errors: {total_err} | Not found: {total_nf}")

if __name__ == '__main__':
    main()
