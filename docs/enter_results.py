#!/usr/bin/env python3
"""Enter February 2026 match scores from Excel into event_id=24 via API."""

import json
import subprocess
import openpyxl

EXCEL = '/Users/Kalibek_Turgumbayev/projects/test/league/docs/Лиги Астана 2026_02_08.xlsx'
COOKIES = '/tmp/admin_cookies.txt'
BASE_URL = 'http://localhost:8080/api/v1'
EVENT_ID = 24

GROUP_IDS = {
    ('S', 1): 96,
    ('A', 1): 97, ('A', 2): 98, ('A', 3): 99, ('A', 4): 100,
    ('A', 5): 101, ('A', 6): 102, ('A', 7): 103, ('A', 8): 104,
    ('B', 1): 105, ('B', 2): 106, ('B', 3): 107, ('B', 4): 108,
    ('B', 5): 109, ('B', 6): 110, ('B', 7): 111, ('B', 8): 112,
    ('C', 1): 113, ('C', 2): 114, ('C', 3): 115, ('C', 4): 116,
    ('C', 5): 117, ('C', 6): 118, ('C', 7): 119, ('C', 8): 120,
    ('D', 1): 121, ('D', 2): 122, ('D', 3): 123, ('D', 4): 124,
    ('D', 5): 125, ('D', 6): 126, ('D', 7): 127, ('D', 8): 128,
    ('E', 1): 129, ('E', 2): 130, ('E', 3): 131, ('E', 4): 132,
    ('E', 5): 133, ('E', 6): 134, ('E', 7): 135, ('E', 8): 136,
}

CYRILLIC_MAP = {'А': 'A', 'В': 'B', 'С': 'C', 'Д': 'D', 'Е': 'E'}


def normalize_division(ch):
    return CYRILLIC_MAP.get(ch, ch).upper()


def parse_sheet_name(name):
    if name in ('Суперлига', 'Superleague'):
        return ('S', 1)
    base = name
    for suffix in ['_2', '_3']:
        if name.endswith(suffix):
            base = name[:-len(suffix)]
            break
    div = normalize_division(base[0])
    try:
        no = int(base[1:])
    except ValueError:
        return None
    return (div, no)


def load_users_from_api():
    """Returns dict: (last_name_upper, first_name_upper) → user_id"""
    r = subprocess.run(
        ['curl', '-s', f'{BASE_URL}/public/players?limit=1000'],
        capture_output=True, text=True
    )
    players = json.loads(r.stdout)
    users = {}
    for p in players:
        key = (p['lastName'].upper(), p['firstName'].upper())
        users[key] = p['userId']
    return users


def find_uid(full_name, users):
    full_name = full_name.strip().rstrip('+').strip()
    parts = full_name.split(None, 1)
    if len(parts) == 2:
        last, first = parts[0].upper(), parts[1].upper()
        if (last, first) in users:
            return users[(last, first)]
        if (first, last) in users:
            return users[(first, last)]
    return None


def get_group_data(group_id):
    r = subprocess.run(
        ['curl', '-s', f'{BASE_URL}/public/events/{EVENT_ID}/groups/{group_id}'],
        capture_output=True, text=True
    )
    return json.loads(r.stdout)


def put_score(group_id, match_id, score1, score2):
    r = subprocess.run(
        ['curl', '-s', '-b', COOKIES, '-X', 'PUT',
         f'{BASE_URL}/secured/groups/{group_id}/matches/{match_id}',
         '-H', 'Content-Type: application/json',
         '-d', json.dumps({'score1': score1, 'score2': score2})],
        capture_output=True, text=True
    )
    return r.stdout.strip()


def parse_score(score_str):
    """Parse '3:1' → (3, 1). Returns None if unparseable."""
    if not score_str or not isinstance(score_str, str):
        return None
    s = score_str.strip()
    if ':' not in s:
        return None
    parts = s.split(':')
    if len(parts) != 2:
        return None
    # Skip W/L forfeit entries
    if parts[0].isalpha() or parts[1].isalpha():
        return None
    try:
        return int(parts[0]), int(parts[1])
    except ValueError:
        return None


def extract_matches(ws):
    """Extract (player1_name, player2_name, score1, score2) from match rows.
    Match rows: col J (idx 9) = p1, col K (idx 10) = '-', col L (idx 11) = p2,
                col P (idx 15) = score.
    Returns list of tuples; score1/score2 are None if match not played."""
    matches = []
    for row in ws.iter_rows(min_row=30, max_row=100, values_only=True):
        p1 = row[9]    # col J
        sep = row[10]  # col K
        p2 = row[11]   # col L
        score_raw = row[15]  # col P
        if not (isinstance(p1, str) and sep == '-' and isinstance(p2, str)):
            continue
        p1 = p1.strip().rstrip('+').strip()
        p2 = p2.strip().rstrip('+').strip()
        score = parse_score(score_raw)
        s1, s2 = (score[0], score[1]) if score else (None, None)
        matches.append((p1, p2, s1, s2))
    return matches


def process_group(sheet_name, group_key, users):
    group_id = GROUP_IDS.get(group_key)
    if group_id is None:
        return None

    # Get API data
    data = get_group_data(group_id)
    players = data.get('players', [])
    api_matches = data.get('matches', [])

    # Build userId → groupPlayerId
    uid_to_gpid = {p['userId']: p['groupPlayerId'] for p in players}

    # Build {frozenset(gp1,gp2) → (matchId, gp1, gp2)}
    match_lookup = {}
    for m in api_matches:
        gp1, gp2 = m['groupPlayer1Id'], m['groupPlayer2Id']
        key = frozenset([gp1, gp2])
        match_lookup[key] = (m['matchId'], gp1, gp2)

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb[sheet_name]
    excel_matches = extract_matches(ws)

    div, no = group_key
    print(f"\n--- {div}{no} (groupId={group_id}, sheet={sheet_name}) ---")

    ok_count = no_score = not_found = error_count = 0

    for p1_name, p2_name, s1, s2 in excel_matches:
        uid1 = find_uid(p1_name, users)
        uid2 = find_uid(p2_name, users)

        if uid1 is None or uid2 is None:
            missing = p1_name if uid1 is None else p2_name
            print(f"  USER NOT FOUND: {missing!r}")
            not_found += 1
            continue

        gp1 = uid_to_gpid.get(uid1)
        gp2 = uid_to_gpid.get(uid2)

        if gp1 is None or gp2 is None:
            missing = p1_name if gp1 is None else p2_name
            print(f"  NOT IN GROUP: {missing!r} (uid={uid1 if gp1 is None else uid2})")
            not_found += 1
            continue

        if s1 is None:
            no_score += 1
            continue

        key = frozenset([gp1, gp2])
        if key not in match_lookup:
            print(f"  MATCH NOT FOUND: {p1_name} vs {p2_name}")
            not_found += 1
            continue

        match_id, api_gp1, api_gp2 = match_lookup[key]
        # Align score to API player order
        if api_gp1 == gp1:
            final_s1, final_s2 = s1, s2
        else:
            final_s1, final_s2 = s2, s1

        resp = put_score(group_id, match_id, final_s1, final_s2)
        try:
            data_r = json.loads(resp)
            if data_r.get('ok'):
                ok_count += 1
            else:
                print(f"  ERROR: {p1_name} vs {p2_name} {s1}:{s2} → {resp}")
                error_count += 1
        except json.JSONDecodeError:
            print(f"  RAW: {p1_name} vs {p2_name}: {resp}")
            error_count += 1

    print(f"  Scored: {ok_count} | Skipped(no score): {no_score} | Errors: {error_count} | Not found: {not_found}")
    return ok_count, no_score, error_count, not_found


def main():
    users = load_users_from_api()
    print(f"Loaded {len(users)} users from API")

    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    # Select canonical sheet per group key (prefer sheet with data)
    canonical = {}
    for sname in wb.sheetnames:
        key = parse_sheet_name(sname)
        if not key or key not in GROUP_IDS:
            continue
        ws = wb[sname]
        # Check for match data
        has_matches = any(
            isinstance(row[9], str) and row[10] == '-' and isinstance(row[11], str)
            for row in ws.iter_rows(min_row=30, max_row=100, values_only=True)
        )
        if has_matches:
            # Count matches
            match_count = sum(
                1 for row in ws.iter_rows(min_row=30, max_row=100, values_only=True)
                if isinstance(row[9], str) and row[10] == '-' and isinstance(row[11], str)
                   and row[15] is not None
            )
            if key not in canonical or match_count > canonical[key][1]:
                canonical[key] = (sname, match_count)

    print(f"Found {len(canonical)} groups with match data")

    total_ok = total_skip = total_err = total_nf = 0
    for key in sorted(canonical.keys()):
        sname, _ = canonical[key]
        result = process_group(sname, key, users)
        if result:
            ok, skip, err, nf = result
            total_ok += ok
            total_skip += skip
            total_err += err
            total_nf += nf

    print(f"\n=== TOTAL SUMMARY ===")
    print(f"Scored: {total_ok} | Skipped(no score/DNS): {total_skip} | Errors: {total_err} | Not found: {total_nf}")


if __name__ == '__main__':
    main()
