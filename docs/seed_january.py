#!/usr/bin/env python3
"""Seed January 2026 league players into event_id=23 groups via curl."""

import csv
import json
import subprocess
import openpyxl

EXCEL = '/Users/Kalibek_Turgumbayev/projects/test/league/docs/Лиги Астана 2026_01_11.xlsx'
USERS_CSV = '/tmp/users.csv'
COOKIES = '/tmp/admin_cookies.txt'
BASE_URL = 'http://localhost:8080/api/v1/secured/events/23/groups'

# Map (division, groupNo) → groupId from event 23
# B1 has duplicate: 54 (wrong) and 56 (correct) — use 56
GROUP_IDS = {
    ('S', 1): 45,
    ('A', 1): 46, ('A', 2): 47, ('A', 3): 48, ('A', 4): 49,
    ('A', 5): 50, ('A', 6): 51, ('A', 7): 52, ('A', 8): 53,
    ('B', 1): 56, ('B', 2): 57, ('B', 3): 58, ('B', 4): 59,
    ('B', 5): 60, ('B', 6): 61, ('B', 7): 62, ('B', 8): 63,
    ('C', 1): 64, ('C', 2): 65, ('C', 3): 66, ('C', 4): 67,
    ('C', 5): 68, ('C', 6): 69, ('C', 7): 70, ('C', 8): 71,
    ('D', 1): 72, ('D', 2): 73, ('D', 3): 74, ('D', 4): 75,
    ('D', 5): 76, ('D', 6): 77, ('D', 7): 78, ('D', 8): 79,
    ('E', 1): 80, ('E', 2): 81, ('E', 3): 82, ('E', 4): 83,
    ('E', 5): 84, ('E', 6): 85, ('E', 7): 86, ('E', 8): 87,
}

# Cyrillic to Latin normalization for division letters
CYRILLIC_MAP = {
    'А': 'A',  # А
    'В': 'B',  # В
    'С': 'C',  # С
    'Д': 'D',  # Д (just in case)
    'Е': 'E',  # Е
}

def normalize_division(ch):
    return CYRILLIC_MAP.get(ch, ch).upper()

def parse_sheet_name(name):
    """Return (division, groupNo) or None if should skip."""
    if name in ('Суперлига', 'Superleague'):
        return ('S', 1)
    # Strip _2 suffix (some groups store real data in _2 sheets)
    base = name.replace('_2', '') if name.endswith('_2') else name
    # First char is division letter (may be Cyrillic)
    div = normalize_division(base[0])
    try:
        no = int(base[1:])
    except ValueError:
        return None
    return (div, no)

def load_users():
    """Returns dict: (last_name_upper, first_name_upper) → user_id"""
    users = {}
    with open(USERS_CSV, newline='', encoding='utf-8') as f:
        for row in csv.reader(f):
            if len(row) < 3:
                continue
            uid, first, last = row[0].strip(), row[1].strip(), row[2].strip()
            try:
                uid = int(uid)
            except ValueError:
                continue
            key = (last.upper(), first.upper())
            users[key] = uid
    return users

def extract_players_from_sheet(ws):
    """Extract player names from rows 7-14 col D (index 3)."""
    players = []
    for row in ws.iter_rows(min_row=7, max_row=14, min_col=4, max_col=4, values_only=True):
        name = row[0]
        if name and isinstance(name, str):
            players.append(name.strip())
    return players

def find_user_id(full_name, users):
    """full_name is 'Фамилия Имя' format."""
    parts = full_name.split(None, 1)
    if len(parts) == 2:
        last, first = parts[0].upper(), parts[1].upper()
        uid = users.get((last, first))
        if uid:
            return uid
        # Try swapped (some sheets may have 'Имя Фамилия')
        uid = users.get((first, last))
        if uid:
            return uid
    # Try full name as last_name with empty first
    uid = users.get((full_name.upper(), ''))
    return uid

def seed_player(group_id, user_id):
    result = subprocess.run(
        ['curl', '-s', '-b', COOKIES, '-X', 'POST',
         f'{BASE_URL}/{group_id}/seed',
         '-H', 'Content-Type: application/json',
         '-d', json.dumps({'userId': user_id})],
        capture_output=True, text=True
    )
    return result.stdout.strip()

def main():
    users = load_users()
    print(f"Loaded {len(users)} users")

    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    total_seeded = 0
    not_found = []
    already_in = []
    errors = []

    for sheet_name in wb.sheetnames:
        group_key = parse_sheet_name(sheet_name)
        if group_key is None:
            continue
        group_id = GROUP_IDS.get(group_key)
        if group_id is None:
            continue

        ws = wb[sheet_name]
        players = extract_players_from_sheet(ws)
        if not any(players):
            continue

        div, no = group_key
        print(f"\n--- {div}{no} (groupId={group_id}, sheet={sheet_name}) ---")

        for name in players:
            uid = find_user_id(name, users)
            if uid is None:
                print(f"  NOT FOUND: {name}")
                not_found.append((sheet_name, name))
                continue

            resp = seed_player(group_id, uid)
            try:
                data = json.loads(resp)
                if data.get('ok'):
                    print(f"  OK: {name} (uid={uid})")
                    total_seeded += 1
                elif 'already' in resp.lower() or 'duplicate' in resp.lower() or 'conflict' in resp.lower():
                    print(f"  ALREADY: {name} (uid={uid}): {resp}")
                    already_in.append((sheet_name, name))
                else:
                    print(f"  ERROR: {name} (uid={uid}): {resp}")
                    errors.append((sheet_name, name, resp))
            except json.JSONDecodeError:
                print(f"  RAW: {name} (uid={uid}): {resp}")
                if resp and resp not in ('', 'null'):
                    errors.append((sheet_name, name, resp))

    print(f"\n=== SUMMARY ===")
    print(f"Seeded: {total_seeded}")
    print(f"Not found: {len(not_found)}")
    if not_found:
        for s, n in not_found:
            print(f"  [{s}] {n}")
    print(f"Already in group: {len(already_in)}")
    print(f"Errors: {len(errors)}")
    if errors:
        for s, n, r in errors:
            print(f"  [{s}] {n}: {r}")

if __name__ == '__main__':
    main()
