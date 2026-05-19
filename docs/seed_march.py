#!/usr/bin/env python3
"""Seed March 2026 (event 25) groups from Excel. Fixes mismatches: removes wrong players, adds missing ones.
Two-pass: remove all wrong players first, then add missing ones."""

import json
import re
import subprocess
import urllib.parse
import openpyxl

EXCEL = '/Users/Kalibek_Turgumbayev/projects/test/league/docs/Лиги Астана 2026_03_22.xlsx'
COOKIES = '/tmp/league_cookies.txt'
BASE_URL = 'http://localhost:8080/api/v1'
EVENT_ID = 25

GROUP_IDS = {
    ('S', 1): 137,
    ('A', 1): 138, ('A', 2): 139, ('A', 3): 140, ('A', 4): 141,
    ('A', 5): 142, ('A', 6): 143, ('A', 7): 144, ('A', 8): 145,
    ('B', 1): 146, ('B', 2): 147, ('B', 3): 148, ('B', 4): 149,
    ('B', 5): 150, ('B', 6): 151, ('B', 7): 152, ('B', 8): 153,
    ('C', 1): 154, ('C', 2): 155, ('C', 3): 156, ('C', 4): 157,
    ('C', 5): 158, ('C', 6): 159, ('C', 7): 160, ('C', 8): 161,
    ('D', 1): 162, ('D', 2): 163, ('D', 3): 164, ('D', 4): 165,
    ('D', 5): 166, ('D', 6): 167, ('D', 7): 168, ('D', 8): 169,
    ('E', 1): 170, ('E', 2): 171, ('E', 3): 172, ('E', 4): 173,
    ('E', 5): 174, ('E', 6): 175, ('E', 7): 176, ('E', 8): 177,
}

CYRILLIC_MAP = {'А': 'A', 'В': 'B', 'С': 'C', 'Д': 'D', 'Е': 'E'}

TRANSLIT = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
    'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
    'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
    'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
    'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    'ә': 'ae', 'ғ': 'gh', 'қ': 'q', 'ң': 'ng', 'ө': 'oe', 'ұ': 'u',
    'ү': 'ue', 'һ': 'h', 'і': 'i',
}


def transliterate(text):
    result = []
    for ch in text.lower():
        result.append(TRANSLIT.get(ch, re.sub(r'[^a-z0-9]', '', ch)))
    return ''.join(result)


def make_email(first, last):
    f = transliterate(first) or 'player'
    l = transliterate(last) or 'unknown'
    return f"{f}_{l}@tt-league.kz"


def norm_div(ch):
    return CYRILLIC_MAP.get(ch, ch).upper()


def parse_sheet(name):
    if name in ('Суперлига', 'Superleague'):
        return ('S', 1)
    base = name
    for s in ['_2', '_3']:
        if name.endswith(s):
            base = name[:-len(s)]
            break
    div = norm_div(base[0])
    try:
        return (div, int(base[1:]))
    except ValueError:
        return None


def curl(*args):
    r = subprocess.run(['curl', '-s'] + list(args), capture_output=True, text=True)
    return r.stdout.strip()


def get_all_players():
    raw = curl(f'{BASE_URL}/public/players?limit=2000')
    players = json.loads(raw)
    lookup = {}
    for p in players:
        key = (p['lastName'].upper(), p['firstName'].upper())
        lookup[key] = p['userId']
    return lookup


def find_uid(full_name, lookup):
    full_name = full_name.strip().rstrip('+').strip()
    if full_name == '-' or not full_name:
        return None
    parts = full_name.split(None, 1)
    if len(parts) == 2:
        last, first = parts[0].upper(), parts[1].upper()
        uid = lookup.get((last, first))
        if uid:
            return uid
        uid = lookup.get((first, last))
        if uid:
            return uid
    uid = lookup.get((full_name.upper(), ''))
    return uid


def create_player(first, last, lookup):
    email = make_email(first, last)
    body = json.dumps({'firstName': first, 'lastName': last, 'email': email})
    raw = curl('-b', COOKIES, '-X', 'POST',
               f'{BASE_URL}/secured/players',
               '-H', 'Content-Type: application/json',
               '-d', body)
    try:
        data = json.loads(raw)
        uid = data.get('userId')
        if uid:
            lookup[(last.upper(), first.upper())] = uid
            print(f"    CREATED: {last} {first} (uid={uid}, email={email})")
            return uid
        else:
            print(f"    CREATE FAILED: {last} {first}: {raw}")
            return None
    except Exception:
        print(f"    CREATE ERROR: {last} {first}: {raw}")
        return None


def get_or_create_player(full_name, lookup):
    uid = find_uid(full_name, lookup)
    if uid:
        return uid
    # Try search API
    encoded = urllib.parse.quote(full_name)
    raw = curl(f'{BASE_URL}/public/players?q={encoded}&limit=10')
    try:
        results = json.loads(raw)
        for p in results:
            api_name = f"{p['lastName']} {p['firstName']}"
            if api_name.upper() == full_name.upper():
                uid = p['userId']
                lookup[(p['lastName'].upper(), p['firstName'].upper())] = uid
                return uid
    except Exception:
        pass
    # Create new player
    parts = full_name.strip().rstrip('+').strip().split(None, 1)
    if len(parts) == 2:
        last, first = parts[0], parts[1]
    else:
        last, first = parts[0], 'Unknown'
    print(f"    NOT FOUND in API, creating: {last} {first}")
    return create_player(first, last, lookup)


def get_group_players(group_id):
    raw = curl(f'{BASE_URL}/public/events/{EVENT_ID}/groups/{group_id}')
    data = json.loads(raw)
    return data.get('players', [])


def remove_player(group_id, gp_id):
    raw = curl('-b', COOKIES, '-X', 'DELETE',
               f'{BASE_URL}/secured/events/{EVENT_ID}/groups/{group_id}/players/{gp_id}')
    return raw


def seed_player(group_id, user_id):
    body = json.dumps({'userId': user_id})
    raw = curl('-b', COOKIES, '-X', 'POST',
               f'{BASE_URL}/secured/events/{EVENT_ID}/groups/{group_id}/seed',
               '-H', 'Content-Type: application/json',
               '-d', body)
    return raw


def extract_excel_players(ws):
    players = []
    for row in ws.iter_rows(min_row=7, max_row=14, min_col=4, max_col=4, values_only=True):
        name = row[0]
        if name and isinstance(name, str):
            cleaned = name.strip().rstrip('+').strip()
            if cleaned and cleaned != '-':
                players.append(cleaned)
    return players


def pick_canonical_sheets(wb):
    """Pick best sheet per group key: prefer sheet with actual player data in col D rows 7-14."""
    candidates = {}  # key → list of (sname, player_count)
    for sname in wb.sheetnames:
        key = parse_sheet(sname)
        if not key or key not in GROUP_IDS:
            continue
        ws = wb[sname]
        players = extract_excel_players(ws)
        if sname not in candidates:
            candidates.setdefault(key, [])
        candidates[key].append((sname, len(players), players))

    canonical = {}
    for key, options in candidates.items():
        # Sort: prefer more players, then prefer non-_2 (original)
        options.sort(key=lambda x: (-x[1], x[0].endswith('_2'), x[0].endswith('_3')))
        best_sname, best_count, best_players = options[0]
        if best_count > 0:
            canonical[key] = (best_sname, best_players)

    return canonical


def main():
    print("Loading players from API...")
    lookup = get_all_players()
    print(f"  {len(lookup)} players loaded")

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    canonical = pick_canonical_sheets(wb)
    print(f"Found {len(canonical)} groups with players in Excel")

    # Build plan: for each group, compute what needs to change
    plan = {}  # group_key → {group_id, excel_names, excel_uids, api_players, to_remove, to_add}

    print("\n--- Building plan ---")
    for key in sorted(canonical.keys()):
        sname, excel_players = canonical[key]
        group_id = GROUP_IDS[key]
        div, no = key

        api_players = get_group_players(group_id)
        api_by_uid = {p['userId']: p for p in api_players}

        # Resolve Excel UIDs
        excel_uids = []
        for name in excel_players:
            uid = get_or_create_player(name, lookup)
            if uid:
                excel_uids.append(uid)
            else:
                print(f"  {div}{no}: SKIP (no uid): {name}")

        excel_uid_set = set(excel_uids)
        api_uid_set = set(api_by_uid.keys())

        to_remove = [(api_by_uid[uid]['groupPlayerId'], uid, api_by_uid[uid]['user']) for uid in (api_uid_set - excel_uid_set)]
        to_add = [uid for uid in excel_uids if uid not in api_uid_set]

        plan[key] = {
            'group_id': group_id,
            'sname': sname,
            'excel_uids': excel_uids,
            'api_by_uid': api_by_uid,
            'to_remove': to_remove,
            'to_add': to_add,
        }

    # PASS 1: Remove all wrong players
    print("\n=== PASS 1: Removing wrong players ===")
    total_removed = 0
    for key in sorted(plan.keys()):
        p = plan[key]
        div, no = key
        for gp_id, uid, user in p['to_remove']:
            name = f"{user['lastName']} {user['firstName']}"
            resp = remove_player(p['group_id'], gp_id)
            try:
                data = json.loads(resp)
                if data.get('ok'):
                    print(f"  {div}{no}: REMOVED {name} (gpid={gp_id})")
                    total_removed += 1
                else:
                    print(f"  {div}{no}: REMOVE ERROR {name}: {resp}")
            except Exception:
                print(f"  {div}{no}: REMOVE RAW {name}: {resp}")
    print(f"Total removed: {total_removed}")

    # PASS 2: Add missing players
    print("\n=== PASS 2: Seeding missing players ===")
    total_seeded = total_skipped = total_errors = 0
    for key in sorted(plan.keys()):
        p = plan[key]
        div, no = key
        for uid in p['to_add']:
            resp = seed_player(p['group_id'], uid)
            try:
                data = json.loads(resp)
                if data.get('ok') or data.get('groupPlayerId'):
                    print(f"  {div}{no}: SEEDED uid={uid}")
                    total_seeded += 1
                elif 'already' in resp.lower() or 'duplicate' in resp.lower() or 'conflict' in resp.lower():
                    print(f"  {div}{no}: ALREADY uid={uid}: {resp}")
                    total_skipped += 1
                else:
                    print(f"  {div}{no}: SEED ERROR uid={uid}: {resp}")
                    total_errors += 1
            except json.JSONDecodeError:
                print(f"  {div}{no}: SEED RAW uid={uid}: {resp}")
                total_errors += 1
    print(f"Total seeded: {total_seeded} | Skipped: {total_skipped} | Errors: {total_errors}")

    print(f"\n=== SUMMARY ===")
    print(f"Removed: {total_removed} | Seeded: {total_seeded} | Skipped (already in event): {total_skipped} | Errors: {total_errors}")

    # Verify
    print("\n=== VERIFICATION ===")
    mismatches = 0
    for key in sorted(plan.keys()):
        p = plan[key]
        div, no = key
        api_players = get_group_players(p['group_id'])
        api_uid_set = set(pp['userId'] for pp in api_players)
        excel_uid_set = set(p['excel_uids'])
        only_api = api_uid_set - excel_uid_set
        only_excel = excel_uid_set - api_uid_set
        if only_api or only_excel:
            print(f"  {div}{no} STILL MISMATCH: only_api_uids={only_api}, only_excel_uids={only_excel}")
            mismatches += 1
    if mismatches == 0:
        print("  All groups match Excel!")
    else:
        print(f"  {mismatches} groups still have mismatches")


if __name__ == '__main__':
    main()
