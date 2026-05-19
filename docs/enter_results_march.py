#!/usr/bin/env python3
"""Enter March 2026 match scores from Excel into event_id=25 via API."""

import json
import subprocess
import openpyxl

EXCEL = '/Users/Kalibek_Turgumbayev/projects/test/league/docs/Лиги Астана 2026_03_22.xlsx'
COOKIES = '/tmp/league_cookies.txt'
BASE_URL = 'http://localhost:8080/api/v1'
EVENT_ID = 25

GROUP_IDS = {
    ('S', 1): 137,
    ('A', 1): 138, ('A', 2): 139, ('A', 3): 140, ('A', 4): 141,
    ('A', 5): 142, ('A', 6): 143, ('A', 7): 144, ('A', 8): 145,
    ('B', 1): 146, ('B', 2): 147, ('B', 3): 148, ('B', 4): 149,
    ('B', 5): 150, ('B', 6): 151, ('B', 7): 152, ('B', 8): 153,
    ('C', 1): 154, ('C', 2): 155, ('C', 3): 156, ('C', 4): 157,
    ('C', 5): 158, ('C', 6): 159, ('C', 7): 160, ('C', 8): 161,
    ('D', 1): 162, ('D', 2): 163, ('D', 3): 164, ('D', 4): 165,
    ('D', 5): 166, ('D', 6): 167, ('D', 7): 168, ('D', 8): 169,
    ('E', 1): 170, ('E', 2): 171, ('E', 3): 172, ('E', 4): 173,
    ('E', 5): 174, ('E', 6): 175, ('E', 7): 176, ('E', 8): 177,
}

CYRILLIC_MAP = {'А': 'A', 'В': 'B', 'С': 'C', 'Д': 'D', 'Е': 'E'}

# Kazakh special chars → Russian equivalents for fuzzy matching
KAZAKH_TO_RU = str.maketrans('ҚқҢңҮүӨөҒғҰұӘәІіҺһ', 'КкНнУуОоГгУуАаИиХх')


def norm_div(ch):
    return CYRILLIC_MAP.get(ch, ch).upper()


def parse_sheet(name):
    if name in ('Суперлига', 'Superleague'):
        return ('S', 1)
    base = name
    for s in ['_2', '_3']:
        if name.endswith(s):
            base = name[:-len(s)]
            break
    div = norm_div(base[0])
    try:
        return (div, int(base[1:]))
    except ValueError:
        return None


def load_users_from_api():
    r = subprocess.run(
        ['curl', '-s', f'{BASE_URL}/public/players?limit=2000'],
        capture_output=True, text=True
    )
    players = json.loads(r.stdout)
    users = {}
    for p in players:
        key = (p['lastName'].upper(), p['firstName'].upper())
        users[key] = p['userId']
        # Also add normalized (Kazakh→Russian) version
        norm_last = p['lastName'].upper().translate(KAZAKH_TO_RU)
        norm_first = p['firstName'].upper().translate(KAZAKH_TO_RU)
        if (norm_last, norm_first) != (p['lastName'].upper(), p['firstName'].upper()):
            users.setdefault((norm_last, norm_first), p['userId'])
    return users


def find_uid(full_name, users):
    full_name = full_name.strip().rstrip('+').strip()
    if not full_name or full_name == '-':
        return None
    parts = full_name.split(None, 1)
    if len(parts) == 2:
        last, first = parts[0].upper(), parts[1].upper()
        if (last, first) in users:
            return users[(last, first)]
        if (first, last) in users:
            return users[(first, last)]
        # Try with Kazakh→Russian normalization
        norm_last = last.translate(KAZAKH_TO_RU)
        norm_first = first.translate(KAZAKH_TO_RU)
        if (norm_last, norm_first) in users:
            return users[(norm_last, norm_first)]
        if (norm_first, norm_last) in users:
            return users[(norm_first, norm_last)]
    return None


def get_group_data(group_id):
    r = subprocess.run(
        ['curl', '-s', f'{BASE_URL}/public/events/{EVENT_ID}/groups/{group_id}'],
        capture_output=True, text=True
    )
    return json.loads(r.stdout)


def put_score(group_id, match_id, score1, score2):
    r = subprocess.run(
        ['curl', '-s', '-b', COOKIES, '-X', 'PUT',
         f'{BASE_URL}/secured/groups/{group_id}/matches/{match_id}',
         '-H', 'Content-Type: application/json',
         '-d', json.dumps({'score1': score1, 'score2': score2})],
        capture_output=True, text=True
    )
    return r.stdout.strip()


def parse_score(score_raw):
    if not score_raw or not isinstance(score_raw, str):
        return None
    s = score_raw.strip()
    if ':' not in s:
        return None
    parts = s.split(':')
    if len(parts) != 2:
        return None
    if parts[0].isalpha() or parts[1].isalpha():
        return None
    try:
        return int(parts[0]), int(parts[1])
    except ValueError:
        return None


def extract_matches(ws):
    matches = []
    for row in ws.iter_rows(min_row=30, max_row=120, values_only=True):
        p1 = row[9]
        sep = row[10]
        p2 = row[11]
        score_raw = row[15]
        if not (isinstance(p1, str) and sep == '-' and isinstance(p2, str)):
            continue
        p1 = p1.strip().rstrip('+').strip()
        p2 = p2.strip().rstrip('+').strip()
        score = parse_score(score_raw)
        s1, s2 = (score[0], score[1]) if score else (None, None)
        matches.append((p1, p2, s1, s2))
    return matches


def extract_players_col_d(ws):
    """Count players in col D rows 7-14 (to detect data sheets vs blank templates)."""
    count = 0
    for row in ws.iter_rows(min_row=7, max_row=14, min_col=4, max_col=4, values_only=True):
        if row[0] and isinstance(row[0], str) and row[0].strip():
            count += 1
    return count


def pick_canonical_sheets(wb):
    """Prefer sheet with actual player data; if tied, prefer sheet with most match data."""
    candidates = {}
    for sname in wb.sheetnames:
        key = parse_sheet(sname)
        if not key or key not in GROUP_IDS:
            continue
        ws = wb[sname]
        player_count = extract_players_col_d(ws)
        match_count = sum(
            1 for row in ws.iter_rows(min_row=30, max_row=120, values_only=True)
            if isinstance(row[9], str) and row[10] == '-' and isinstance(row[11], str)
        )
        scored_count = sum(
            1 for row in ws.iter_rows(min_row=30, max_row=120, values_only=True)
            if isinstance(row[9], str) and row[10] == '-' and isinstance(row[11], str)
               and row[15] is not None
        )
        candidates.setdefault(key, []).append((sname, player_count, match_count, scored_count))

    canonical = {}
    for key, options in candidates.items():
        # Prefer: more players, then more scored matches, then more total matches
        options.sort(key=lambda x: (-x[1], -x[3], -x[2]))
        best = options[0]
        if best[2] > 0:  # has any matches at all
            canonical[key] = best[0]
    return canonical


def process_group(sheet_name, group_key, users):
    group_id = GROUP_IDS[group_key]
    data = get_group_data(group_id)
    players = data.get('players', [])
    api_matches = data.get('matches', [])

    uid_to_gpid = {p['userId']: p['groupPlayerId'] for p in players}

    match_lookup = {}
    for m in api_matches:
        gp1, gp2 = m['groupPlayer1Id'], m['groupPlayer2Id']
        key = frozenset([gp1, gp2])
        match_lookup[key] = (m['matchId'], gp1, gp2)

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb[sheet_name]
    excel_matches = extract_matches(ws)

    div, no = group_key
    print(f"\n--- {div}{no} (groupId={group_id}, sheet={sheet_name}) ---")

    ok_count = no_score = not_found = error_count = 0

    for p1_name, p2_name, s1, s2 in excel_matches:
        uid1 = find_uid(p1_name, users)
        uid2 = find_uid(p2_name, users)

        if uid1 is None or uid2 is None:
            missing = p1_name if uid1 is None else p2_name
            print(f"  USER NOT FOUND: {missing!r}")
            not_found += 1
            continue

        gp1 = uid_to_gpid.get(uid1)
        gp2 = uid_to_gpid.get(uid2)

        if gp1 is None or gp2 is None:
            missing = p1_name if gp1 is None else p2_name
            print(f"  NOT IN GROUP: {missing!r} (uid={uid1 if gp1 is None else uid2})")
            not_found += 1
            continue

        if s1 is None:
            no_score += 1
            continue

        key = frozenset([gp1, gp2])
        if key not in match_lookup:
            print(f"  MATCH NOT FOUND: {p1_name} vs {p2_name}")
            not_found += 1
            continue

        match_id, api_gp1, api_gp2 = match_lookup[key]
        if api_gp1 == gp1:
            final_s1, final_s2 = s1, s2
        else:
            final_s1, final_s2 = s2, s1

        resp = put_score(group_id, match_id, final_s1, final_s2)
        try:
            data_r = json.loads(resp)
            if data_r.get('ok'):
                ok_count += 1
            else:
                print(f"  ERROR: {p1_name} vs {p2_name} {s1}:{s2} → {resp}")
                error_count += 1
        except json.JSONDecodeError:
            print(f"  RAW: {p1_name} vs {p2_name}: {resp}")
            error_count += 1

    print(f"  Scored: {ok_count} | Skipped(no score): {no_score} | Errors: {error_count} | Not found: {not_found}")
    return ok_count, no_score, error_count, not_found


def main():
    users = load_users_from_api()
    print(f"Loaded {len(users)} users from API")

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    canonical = pick_canonical_sheets(wb)
    print(f"Found {len(canonical)} groups with match data")

    total_ok = total_skip = total_err = total_nf = 0
    for key in sorted(canonical.keys()):
        sname = canonical[key]
        result = process_group(sname, key, users)
        if result:
            ok, skip, err, nf = result
            total_ok += ok
            total_skip += skip
            total_err += err
            total_nf += nf

    print(f"\n=== TOTAL SUMMARY ===")
    print(f"Scored: {total_ok} | Skipped(no score/DNS): {total_skip} | Errors: {total_err} | Not found: {total_nf}")


if __name__ == '__main__':
    main()
